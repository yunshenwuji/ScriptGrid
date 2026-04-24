# -*- coding: utf-8 -*-
"""
述格前端测试 - 输出文件对比工具
用于对比浏览器下载的转换结果与预期输出文件
"""

import sys
import os

def compare_srt(file1, file2):
    """对比两个SRT文件，返回差异列表"""
    diffs = []
    
    with open(file1, 'r', encoding='utf-8') as f:
        lines1 = f.readlines()
    with open(file2, 'r', encoding='utf-8') as f:
        lines2 = f.readlines()
    
    max_lines = max(len(lines1), len(lines2))
    
    for i in range(max_lines):
        line1 = lines1[i].rstrip('\n') if i < len(lines1) else '<文件结束>'
        line2 = lines2[i].rstrip('\n') if i < len(lines2) else '<文件结束>'
        
        if line1 != line2:
            diffs.append({
                'line': i + 1,
                'actual': line1,
                'expected': line2
            })
    
    # 汇总信息
    result = {
        'match': len(diffs) == 0,
        'actual_lines': len(lines1),
        'expected_lines': len(lines2),
        'diff_count': len(diffs),
        'diffs': diffs[:50]  # 最多显示50处差异
    }
    return result


def compare_xlsx(file1, file2):
    """对比两个XLSX文件，返回差异列表"""
    from openpyxl import load_workbook
    
    diffs = []
    
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)
    
    # 对比Sheet数量
    if len(wb1.sheetnames) != len(wb2.sheetnames):
        diffs.append({
            'type': 'sheet_count',
            'actual': len(wb1.sheetnames),
            'expected': len(wb2.sheetnames)
        })
    
    # 对比每个Sheet
    for sheet_name in wb2.sheetnames:
        if sheet_name not in wb1.sheetnames:
            diffs.append({
                'type': 'missing_sheet',
                'sheet': sheet_name
            })
            continue
        
        ws1 = wb1[sheet_name]
        ws2 = wb2[sheet_name]
        
        max_row = max(ws1.max_row, ws2.max_row)
        max_col = max(ws1.max_column, ws2.max_column)
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                val1 = ws1.cell(row=row, column=col).value
                val2 = ws2.cell(row=row, column=col).value
                
                if val1 != val2:
                    diffs.append({
                        'type': 'cell_mismatch',
                        'sheet': sheet_name,
                        'cell': f'{ws1.cell(row=row, column=col).coordinate}',
                        'actual': str(val1) if val1 is not None else '<空>',
                        'expected': str(val2) if val2 is not None else '<空>'
                    })
    
    result = {
        'match': len(diffs) == 0,
        'diff_count': len(diffs),
        'diffs': diffs[:50]
    }
    return result


def compare_files(actual, expected):
    """根据文件扩展名自动选择对比方式"""
    if not os.path.exists(actual):
        return {'match': False, 'error': f'实际输出文件不存在: {actual}'}
    if not os.path.exists(expected):
        return {'match': False, 'error': f'预期输出文件不存在: {expected}'}
    
    ext = os.path.splitext(expected)[1].lower()
    
    if ext == '.srt':
        return compare_srt(actual, expected)
    elif ext == '.xlsx':
        return compare_xlsx(actual, expected)
    else:
        return {'match': False, 'error': f'不支持的文件格式: {ext}'}


def print_result(test_name, result, quiet=False):
    """打印对比结果"""
    if quiet:
        # 精简模式：仅输出一行关键信息，便于脚本解析
        if 'error' in result:
            print(f"FAIL|{result['error']}")
            return
        if result['match']:
            print('PASS|files match')
        else:
            diff_count = result.get('diff_count', 0)
            actual_lines = result.get('actual_lines', '?')
            expected_lines = result.get('expected_lines', '?')
            if actual_lines != '?':
                print(f'FAIL|{diff_count} diffs (actual={actual_lines} lines, expected={expected_lines} lines)')
            else:
                print(f'FAIL|{diff_count} diffs')
        return

    print(f"\n{'='*60}")
    print(f"测试: {test_name}")
    print(f"{'='*60}")
    
    if 'error' in result:
        print(f"❌ 错误: {result['error']}")
        return
    
    if result['match']:
        print("✅ 文件内容完全一致")
    else:
        print(f"❌ 文件内容存在差异 (共 {result['diff_count']} 处)")
        
        if 'actual_lines' in result:
            print(f"   实际行数: {result['actual_lines']}, 预期行数: {result['expected_lines']}")
        
        for d in result['diffs'][:20]:
            if 'line' in d:
                print(f"   行 {d['line']}:")
                print(f"     实际: {d['actual'][:100]}")
                print(f"     预期: {d['expected'][:100]}")
            elif d['type'] == 'cell_mismatch':
                print(f"   单元格 {d['sheet']}!{d['cell']}:")
                print(f"     实际: {d['actual'][:100]}")
                print(f"     预期: {d['expected'][:100]}")
            else:
                print(f"   {d}")


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Compare test output files')
    parser.add_argument('actual', help='Actual output file path')
    parser.add_argument('expected', help='Expected output file path')
    parser.add_argument('--quiet', '-q', action='store_true', help='Quiet mode: output one-line summary')
    args = parser.parse_args()

    result = compare_files(args.actual, args.expected)
    print_result('文件对比', result, quiet=args.quiet)
