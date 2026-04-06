"""
XLSX 字幕表格解析器
负责从 .xlsx 文件读取字幕数据。
"""

import logging
from openpyxl import load_workbook
from exceptions import ParseError
import constants

# 配置日志记录器
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
if not logger.handlers:
    handler = logging.StreamHandler()
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)
    logger.addHandler(handler)


def parse_xlsx(file_path):
    """
    解析 .xlsx 字幕表格文件。
    :param file_path: .xlsx 文件的路径。
    :return: 一个二维列表,每个子列表代表一行字幕:[序号, 开始时间, 结束时间, 字幕内容]。
    :raises ParseError: 当解析过程出错时(例如表头不正确)。
    """
    try:
        # 加载工作簿和活动工作表
        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active

        # 检查表头(只检查前4列,忽略后面可能存在的空列)
        header_row = [cell.value for cell in next(ws.iter_rows())]
        expected_header = constants.EXCEL_HEADERS
        expected_len = len(expected_header)
        # 确保至少有4列
        if len(header_row) < expected_len:
            raise ParseError(constants.MSG_WARNING_INCORRECT_HEADER.format(expected=expected_header, actual=header_row))
        # 只比较前4列,忽略后面的冗余列
        if header_row[:expected_len] != expected_header:
            raise ParseError(constants.MSG_WARNING_INCORRECT_HEADER.format(expected=expected_header, actual=header_row[:expected_len]))

        data = []
        # 从第二行开始迭代数据行
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 检查行是否为空或不完整
            if not any(cell is not None for cell in row):
                continue # 跳过空行
            if len(row) < 4:
                logger.warning(f"发现不完整的数据行: {row}")
                continue # 跳过不完整的行

            index, start_time, end_time, text = row[0], row[1], row[2], row[3]
            # 确保所有字段都转换为字符串,处理可能的数字类型
            data.append([str(index), str(start_time), str(end_time), str(text) if text else ""])

        return data

    except ParseError:
        # 重新抛出我们自定义的 ParseError
        raise
    except Exception as e:
        raise ParseError(f"解析 Excel 文件 '{file_path}' 时出错: {e}") from e
