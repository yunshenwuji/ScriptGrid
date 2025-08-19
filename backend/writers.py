"""
文件写入和读取模块
负责将内部数据结构写入 .xlsx 或 .srt 文件，以及从 .xlsx 文件读取数据。
"""

import os
from openpyxl import Workbook, load_workbook
from exceptions import WriteError, ParseError
import constants

def write_to_excel(data, output_path):
    """
    将提取的数据写入一个 .xlsx 文件。
    :param data: 包含所有字幕信息的二维列表。
    :param output_path: 输出的 .xlsx 文件的完整路径。
    :raises WriteError: 当写入过程出错时。
    """
    try:
        wb = Workbook() # 创建一个新的Excel工作簿
        ws = wb.active  # 获取当前活动的工作表
        ws.title = constants.EXCEL_SHEET_NAME # 根据常量设置工作表名称

        # 写入表头
        ws.append(constants.EXCEL_HEADERS) # append 方法可以直接写入一行

        # 遍历数据，将每一行字幕写入Excel工作表
        for row_data in data:
            ws.append(row_data)

        # 保存工作簿到指定的路径，如果文件已存在则会覆盖
        wb.save(output_path)
    except Exception as e:
        raise WriteError(f"写入 Excel 文件 '{output_path}' 时出错: {e}") from e


def write_to_srt(data, output_path):
    """
    将提取的数据写入一个 .srt 文件。
    :param data: 包含所有字幕信息的二维列表。格式: [序号, 开始时间, 结束时间, 字幕内容]
    :param output_path: 输出的 .srt 文件的完整路径。
    :raises WriteError: 当写入过程出错时。
    """
    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            for i, entry in enumerate(data):
                index, start_time, end_time, text = entry
                # SRT格式要求：序号、时间码、文本、空行
                f.write(f"{index}\n")
                f.write(f"{start_time} --> {end_time}\n")
                f.write(f"{text}\n")
                f.write("\n") # 块之间的空行
    except Exception as e:
        raise WriteError(f"写入 SRT 文件 '{output_path}' 时出错: {e}") from e


def parse_xlsx(file_path):
    """
    解析 .xlsx 字幕表格文件。
    :param file_path: .xlsx 文件的路径。
    :return: 一个二维列表，每个子列表代表一行字幕：[序号, 开始时间, 结束时间, 字幕内容]。
    :raises ParseError: 当解析过程出错时（例如表头不正确）。
    """
    try:
        # 加载工作簿和活动工作表
        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active

        # 检查表头
        header_row = [cell.value for cell in next(ws.iter_rows())]
        expected_header = constants.EXCEL_HEADERS
        if header_row != expected_header:
            raise ParseError(constants.MSG_WARNING_INCORRECT_HEADER.format(expected=expected_header, actual=header_row))

        data = []
        # 从第二行开始迭代数据行
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 检查行是否为空或不完整
            if not any(cell is not None for cell in row):
                continue # 跳过空行
            if len(row) < 4:
                print(f"警告: 发现不完整的数据行: {row}")
                continue # 跳过不完整的行

            index, start_time, end_time, text = row[0], row[1], row[2], row[3]
            # 确保所有字段都转换为字符串，处理可能的数字类型
            data.append([str(index), str(start_time), str(end_time), str(text) if text else ""])

        return data

    except ParseError:
        # 重新抛出我们自定义的 ParseError
        raise
    except Exception as e:
        raise ParseError(f"解析 Excel 文件 '{file_path}' 时出错: {e}") from e